#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
채용공고 크롤러 — 원티드/사람인/잡코리아 수집 + 필터 + 리포트 저장
발송/스냅샷은 main.py 에서 담당
"""

import json
import re
import time
from pathlib import Path
from urllib.parse import quote

# 경력 4년 이상 필수 공고 차단 패턴 모음 ("3년 이상"은 통과, "4년~"부터 차단)
_HIGH_CAREER_PATTERNS = [
    # "4년 이상", "4년+", "4년↑", "4년~", "4년차 이상"
    re.compile(r"([4-9]|1\d)\s*년\s*(?:차\s*)?(?:이상|\+|↑|~)"),
    # 범위 표기: "4~10년", "5-10년", "경력 5~7년"
    re.compile(r"([4-9]|1\d)\s*[~\-]\s*\d+\s*년"),
    # "4년차" 단독 (e.g., "4년차 백엔드"). "4년차 이하/미만"은 제외
    re.compile(r"([4-9]|1\d)\s*년\s*차(?!\s*(?:이하|미만))"),
    # "경력 4년" 평문 (이하/미만 제외)
    re.compile(r"경력\s*([4-9]|1\d)\s*년(?!\s*(?:이하|미만|\d))"),
]


def _has_high_career(text: str) -> bool:
    return any(p.search(text) for p in _HIGH_CAREER_PATTERNS)

import requests
from bs4 import BeautifulSoup

# ===== 설정 =====
BASE_DIR = Path(__file__).parent
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)
CONFIG_PATH = BASE_DIR / "config.json"

DEFAULT_CONFIG = {
    "keywords": ["백엔드", "서버", "풀스택", "Python", "Java", "Node.js", "Spring", "FastAPI"],
    "exclude_keywords": ["PHP 전문", "경력 5년 이상 필수", "시니어"],
    "years_min": 1,
    "years_max": 3,
    "include_newbie": True,
    "locations": ["서울", "경기", "인천", "수도권", "원격"],
    "sites": {
        "wanted": True,
        "saramin": True,
        "jobkorea": True
    },
    "max_per_site": 20,
    "discord": {
        "webhook_url": "",
        "bot_token": "",
        "channel_id": "",
        "top_n": 10
    }
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        example = CONFIG_PATH.parent / "config.example.json"
        hint = f" config.example.json을 복사해서 설정하세요: cp {example} {CONFIG_PATH}" if example.exists() else ""
        raise FileNotFoundError(f"config.json이 없습니다.{hint}")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    # 누락 키 기본값으로 보충
    for k, v in DEFAULT_CONFIG.items():
        cfg.setdefault(k, v)
    _validate_config(cfg)
    return cfg


def _validate_config(cfg: dict):
    """필수 항목 누락/타입 오류를 시작 시점에 바로 감지."""
    if not isinstance(cfg.get("keywords"), list) or not cfg["keywords"]:
        raise ValueError("config.json: 'keywords' 는 비어 있지 않은 리스트여야 합니다.")
    if not isinstance(cfg.get("sites"), dict):
        raise ValueError("config.json: 'sites' 는 dict 여야 합니다.")
    discord = cfg.get("discord", {})
    if not isinstance(discord, dict):
        raise ValueError("config.json: 'discord' 는 dict 여야 합니다.")
    top_n = discord.get("top_n", 10)
    try:
        int(top_n)
    except (TypeError, ValueError):
        raise ValueError(f"config.json: 'discord.top_n' 는 정수여야 합니다. (현재: {top_n!r})")


# ===== 원티드 =====
WANTED_DEV_PARENT_ID = 518   # 개발 직군 parent_id
WANTED_LOC = {"서울", "경기", "인천"}

def fetch_wanted(keyword: str = "", limit: int = 50) -> list:
    """
    navigation API로 전체 공고 수집 후 클라이언트 필터링.
    - 개발 직군 (parent_id=518)
    - 경력 1~3년 (annual_from <= 3)
    - 서울/경기/인천 또는 원격
    keyword 파라미터는 호환성 유지용 (실제 필터링은 title 기반)
    """
    results = []
    url = "https://www.wanted.co.kr/api/chaos/navigation/v1/results"
    HEADERS_W = {**HEADERS, "Referer": "https://www.wanted.co.kr/"}
    try:
        raw = []
        for offset in range(0, 301, 100):
            params = {
                "country": "kr",
                "job_sort": "job.latest_order",
                "years": "-1",
                "locations": "all",
                "limit": "100",
                "offset": str(offset),
            }
            r = requests.get(url, params=params, headers=HEADERS_W, timeout=15)
            if r.status_code != 200:
                break
            data = r.json().get("data", [])
            if not data:
                break
            raw.extend(data)
            time.sleep(0.5)

        for item in raw:
            cat = item.get("category_tag") or {}
            if cat.get("parent_id") != WANTED_DEV_PARENT_ID:
                continue
            if item.get("annual_from", 99) > 3:
                continue
            addr = item.get("address") or {}
            loc = addr.get("location", "")
            if loc and not any(l in loc for l in WANTED_LOC):
                continue
            pid = item.get("id")
            raw_tags = item.get("skill_tags") or []
            skills = [t.get("title", "") for t in raw_tags if isinstance(t, dict) and t.get("title")]
            results.append({
                "site": "원티드",
                "title": item.get("position", ""),
                "company": (item.get("company") or {}).get("name", ""),
                "location": loc,
                "experience": f"{item.get('annual_from', 0)}~{item.get('annual_to', '?')}년",
                "url": f"https://www.wanted.co.kr/wd/{pid}" if pid else "",
                "tags": ", ".join(skills),
            })
        print(f"[wanted] {len(results)}건 수집")
    except Exception as e:
        print(f"[wanted] 오류: {e}")
    return results[:limit]


# ===== 사람인 =====
def fetch_saramin(keyword: str, limit: int = 20) -> list:
    results = []
    try:
        url = "https://www.saramin.co.kr/zf_user/search/recruit"
        params = {
            "searchType": "search",
            "searchword": keyword,
            "loc_mcd": "101000,102000,108000",  # 서울, 경기, 인천
            "exp_cd": "1,2,3",  # 1~3년
            "recruitPage": "1",
            "recruitSort": "reg_dt",
        }
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        items = soup.select(".item_recruit")[:limit]
        for it in items:
            title_a = it.select_one(".area_job .job_tit a")
            company_a = it.select_one(".area_corp .corp_name a")
            cond = it.select_one(".job_condition")
            sector = it.select_one(".job_sector")
            title = title_a.get("title", "").strip() if title_a else ""
            href = title_a.get("href", "") if title_a else ""
            url_full = "https://www.saramin.co.kr" + href if href else ""
            company = company_a.get_text(strip=True) if company_a else ""
            cond_text = cond.get_text(" / ", strip=True) if cond else ""
            sector_text = sector.get_text(" / ", strip=True) if sector else ""
            results.append({
                "site": "사람인",
                "title": title,
                "company": company,
                "location": cond_text,
                "experience": "1~3년",
                "url": url_full,
                "tags": sector_text,
            })
    except Exception as e:
        print(f"[saramin] {keyword} 오류: {e}")
    return results


# ===== 잡코리아 =====
def fetch_jobkorea(keyword: str, limit: int = 20) -> list:
    results = []
    try:
        url = f"https://www.jobkorea.co.kr/Search/?stext={quote(keyword)}&careerType=1&careerMin=1&careerMax=3&local=I000"
        r = requests.get(url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        items = soup.select("ul.clear > li.list-post, li.list-post")[:limit]
        for it in items:
            title_a = it.select_one("a.title")
            company_a = it.select_one("a.name") or it.select_one(".post-list-corp a")
            etc = it.select_one(".option, .etc")
            title = title_a.get_text(strip=True) if title_a else ""
            href = title_a.get("href", "") if title_a else ""
            url_full = "https://www.jobkorea.co.kr" + href if href.startswith("/") else href
            company = company_a.get_text(strip=True) if company_a else ""
            etc_text = etc.get_text(" / ", strip=True) if etc else ""
            if not title:
                continue
            results.append({
                "site": "잡코리아",
                "title": title,
                "company": company,
                "location": etc_text,
                "experience": "1~3년",
                "url": url_full,
                "tags": keyword,
            })
    except Exception as e:
        print(f"[jobkorea] {keyword} 오류: {e}")
    return results


# ===== 필터링 / 중복제거 =====
def matches_filter(job: dict, cfg: dict) -> bool:
    text = " ".join([
        job.get("title", ""), job.get("company", ""),
        job.get("location", ""), job.get("tags", ""),
    ]).lower()
    for ex in cfg.get("exclude_keywords", []):
        if ex.lower() in text:
            return False
    # 경력 4년 이상 필수 공고 제외 (사람인은 location 조건 영역에 경력 표기됨)
    career_text = " ".join([
        job.get("title", ""),
        job.get("location", ""),
        job.get("tags", ""),
        job.get("experience", ""),
    ])
    if _has_high_career(career_text):
        return False
    # 지역 필터 (location 정보가 있을 때만)
    loc = job.get("location", "")
    if loc:
        loc_ok = any(l in loc for l in cfg.get("locations", []))
        if not loc_ok:
            # 원격 가능/전국 등의 표기도 통과
            if not any(k in text for k in ["원격", "재택", "전국"]):
                return False
    return True


def dedupe(jobs: list) -> list:
    seen = set()
    out = []
    for j in jobs:
        key = (j.get("company", "").strip(), j.get("title", "").strip())
        if key in seen or not key[1]:
            continue
        seen.add(key)
        out.append(j)
    return out


# ===== 수집 메인 =====
def collect_all(cfg: dict) -> list:
    all_jobs = []
    keywords = cfg.get("keywords", [])
    limit = int(cfg.get("max_per_site", 20))

    # 원티드: 키워드 무관하게 1회만 전체 수집 후 클라이언트 필터링
    if cfg["sites"].get("wanted"):
        all_jobs.extend(fetch_wanted(limit=50))
        time.sleep(0.8)

    for kw in keywords:
        if cfg["sites"].get("saramin"):
            all_jobs.extend(fetch_saramin(kw, limit=limit))
            time.sleep(0.8)
        if cfg["sites"].get("jobkorea"):
            all_jobs.extend(fetch_jobkorea(kw, limit=limit))
            time.sleep(0.8)

    filtered = [j for j in all_jobs if matches_filter(j, cfg)]
    return dedupe(filtered)


# ===== 리포트 생성 =====
def to_html(jobs: list, today: str) -> str:
    by_site = {}
    for j in jobs:
        by_site.setdefault(j["site"], []).append(j)

    html = [f"""<!doctype html><html><head><meta charset="utf-8"><style>
body {{ font-family: -apple-system, 'Apple SD Gothic Neo', sans-serif; color:#222; }}
h1 {{ color:#1a73e8; }}
h2 {{ border-bottom:2px solid #eee; padding-bottom:4px; margin-top:28px; }}
table {{ border-collapse:collapse; width:100%; margin-top:8px; font-size:14px; }}
th, td {{ border:1px solid #e5e5e5; padding:8px; vertical-align:top; text-align:left; }}
th {{ background:#f5f7fb; }}
tr:nth-child(even) td {{ background:#fafbfc; }}
a {{ color:#1a73e8; text-decoration:none; }}
.small {{ color:#888; font-size:12px; }}
</style></head><body>"""]
    html.append(f"<h1>오늘의 채용공고 ({today})</h1>")
    html.append(f"<p class='small'>백엔드/풀스택 · 주니어(1~3년) · 서울/수도권 · 총 <b>{len(jobs)}</b>건</p>")

    if not jobs:
        html.append("<p>오늘은 조건에 맞는 새 공고가 없어요. 내일 다시 확인할게요.</p>")

    for site, items in by_site.items():
        html.append(f"<h2>{site} ({len(items)}건)</h2>")
        html.append("<table><tr><th>회사</th><th>공고</th><th>지역/조건</th><th>태그</th></tr>")
        for j in items:
            title_link = f"<a href='{j.get('url','')}' target='_blank'>{j.get('title','')}</a>" if j.get("url") else j.get("title", "")
            html.append(
                f"<tr><td>{j.get('company','')}</td>"
                f"<td>{title_link}</td>"
                f"<td>{j.get('location','')}</td>"
                f"<td>{j.get('tags','')}</td></tr>"
            )
        html.append("</table>")
    html.append("</body></html>")
    return "".join(html)


def to_xlsx(jobs: list, path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "채용공고"
    headers = ["사이트", "회사", "공고명", "지역/조건", "경력", "태그", "URL"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A73E8")
        c.alignment = Alignment(horizontal="center", vertical="center")
    widths = [10, 20, 50, 30, 10, 30, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for j in jobs:
        ws.append([
            j.get("site", ""), j.get("company", ""), j.get("title", ""),
            j.get("location", ""), j.get("experience", ""),
            j.get("tags", ""), j.get("url", ""),
        ])
    wb.save(path)


if __name__ == "__main__":
    # 단독 실행 시 크롤링 결과만 출력 (발송은 main.py 에서)
    import datetime
    cfg = load_config()
    today = datetime.date.today().isoformat()
    jobs = collect_all(cfg)
    print(f"[result] {len(jobs)}건 수집")
    xlsx_path = REPORTS_DIR / f"jobs_{today}.xlsx"
    to_xlsx(jobs, xlsx_path)
    html = to_html(jobs, today)
    (REPORTS_DIR / f"jobs_{today}.html").write_text(html, encoding="utf-8")
    print(json.dumps({"count": len(jobs), "xlsx": str(xlsx_path)}, ensure_ascii=False))
