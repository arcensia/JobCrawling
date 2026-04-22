#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
채용공고 수집 봇
- 원티드, 사람인, 잡코리아에서 백엔드/풀스택 주니어 (1-3 년차) 서울/수도권 공고 수집
- HTML 이메일로 정리해 발송 (SMTP 필요)
- 엑셀 리포트 동시 저장
- systemd 기반 자동화 (로그 및 상태 관리 지원)
"""

import os
import sys
import json
import time
import smtplib
import logging
import traceback
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path
from urllib.parse import quote
from datetime import datetime

import requests
from bs4 import BeautifulSoup

# ===== 설정 =====
BASE_DIR = Path(__file__).parent
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)
CONFIG_PATH = BASE_DIR / "config.json"
LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"job_bot_{datetime.now().strftime('%Y%m%d')}.log"

# ===== 로거 설정 =====
logger = logging.getLogger("job_bot")
logger.setLevel(logging.INFO)

# 콘솔 핸들러 (테스트 시 유용)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

# 파일 핸들러 (배포 시 유용)
file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

logger.addHandler(console_handler)
logger.addHandler(file_handler)

DEFAULT_CONFIG = {
    "keywords": ["백엔드", "서버", "풀스택", "Python", "Java", "Node.js", "Spring", "FastAPI"],
    "exclude_keywords": ["PHP 전문", "경력 5 년 이상 필수", "시니어"],
    "years_min": 1,
    "years_max": 3,
    "include_newbie": True,
    "locations": ["서울", "경기", "인천", "수도권", "원격"],
    "sites": {
        "wanted": True,
        "saramin": True,
        "jobkorea": True
    },
    "max_per_site": 20,
    "email": {
        "enabled": False,
        "smtp_server": "smtp.gmail.com",
        "smtp_port": 587,
        "sender": "YOUR_GMAIL@gmail.com",
        "app_password": "YOUR_16_DIGIT_APP_PASSWORD",
        "recipient": "zxcklwe@gmail.com"
    }
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}


def load_config() -> dict:
    """설정 파일 로드 (없으면 기본값 생성)"""
    if not CONFIG_PATH.exists():
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)
        logger.info(f"설정 파일이 없음. 기본값으로 생성: {CONFIG_PATH}")
        return DEFAULT_CONFIG
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    # 누락 키 보충
    for k, v in DEFAULT_CONFIG.items():
        cfg.setdefault(k, v)
    return cfg


def save_status(count: int, message: str):
    """실행 상태 저장 (status.json)"""
    status_file = BASE_DIR / "status.json"
    status = {
        "last_run": datetime.now().isoformat(),
        "job_count": count,
        "status": "success" if count >= 0 else "error",
        "message": message,
    }
    with open(status_file, "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)


# ===== 원티드 =====
def fetch_wanted(keyword: str, limit: int = 20) -> list:
    """원티드 public search API 사용"""
    results = []
    try:
        url = "https://www.wanted.co.kr/api/chaos/jobs/v4/jobs"
        params = {
            "1_10": "",
            "country": "kr",
            "job_sort": "job.latest_order",
            "locations": "seoul.all,gyeonggi.all,incheon.all",
            "years": "1-3",
            "limit": str(limit),
            "offset": "0",
            "query": keyword,
        }
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            # fallback: search page HTML
            return fetch_wanted_html(keyword, limit)
        data = r.json()
        for item in (data.get("data") or [])[:limit]:
            pos = item.get("position") or item.get("title") or ""
            company = (item.get("company") or {}).get("name", "")
            loc = (item.get("address") or {}).get("location", "")
            pid = item.get("id") or item.get("job_id")
            link = f"https://www.wanted.co.kr/wd/{pid}" if pid else ""
            results.append({
                "site": "원티드",
                "title": pos,
                "company": company,
                "location": loc,
                "experience": "1~3 년",
                "url": link,
                "tags": ", ".join([t for t in (item.get("tag_names") or [])]),
            })
    except Exception as e:
        logger.error(f"[wanted] {keyword} 오류: {e}")
    return results


def fetch_wanted_html(keyword: str, limit: int) -> list:
    """원티드 검색 페이지 HTML 파싱 (fallback)"""
    results = []
    try:
        url = f"https://www.wanted.co.kr/search?query={quote(keyword)}&tab=position"
        r = requests.get(url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='/wd/']")[:limit]:
            href = a.get("href", "")
            title = a.get_text(" ", strip=True)[:100]
            if not title or not href:
                continue
            results.append({
                "site": "원티드",
                "title": title,
                "company": "",
                "location": "",
                "experience": "",
                "url": "https://www.wanted.co.kr" + href,
                "tags": keyword,
            })
    except Exception as e:
        logger.error(f"[wanted_html] {keyword} 오류: {e}")
    return results


# ===== 사람인 =====
def fetch_saramin(keyword: str, limit: int = 20) -> list:
    results = []
    try:
        url = "https://www.saramin.co.kr/zf_user/search/recruit"
        params = {
            "searchType": "search",
            "searchword": keyword,
            "loc_mcd": "101000,102000,108000",  # 서울, 경기, 인천
            "exp_cd": "1,2,3",  # 1~3 년
            "recruitPage": "1",
            "recruitSort": "reg_dt",
        }
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        items = soup.select(".item_recruit")[:limit]
        for it in items:
            title_a = it.select_one(".area_job .job_tit a")
            company_a = it.select_one(".area_corp .corp_name a")
            cond = it.select_one(".job_condition")
            sector = it.select_one(".job_sector")
            title = title_a.get("title", "").strip() if title_a else ""
            href = title_a.get("href", "") if title_a else ""
            url_full = "https://www.saramin.co.kr" + href if href else ""
            company = company_a.get_text(strip=True) if company_a else ""
            cond_text = cond.get_text(" / ", strip=True) if cond else ""
            sector_text = sector.get_text(" / ", strip=True) if sector else ""
            results.append({
                "site": "사람인",
                "title": title,
                "company": company,
                "location": cond_text,
                "experience": "1~3 년",
                "url": url_full,
                "tags": sector_text,
            })
    except Exception as e:
        logger.error(f"[saramin] {keyword} 오류: {e}")
    return results


# ===== 잡코리아 =====
def fetch_jobkorea(keyword: str, limit: int = 20) -> list:
    results = []
    try:
        url = f"https://www.jobkorea.co.kr/Search/?stext={quote(keyword)}&careerType=1&careerMin=1&careerMax=3&local=I000"
        r = requests.get(url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        items = soup.select("ul.clear > li.list-post, li.list-post")[:limit]
        for it in items:
            title_a = it.select_one("a.title")
            company_a = it.select_one("a.name") or it.select_one(".post-list-corp a")
            etc = it.select_one(".option, .etc")
            title = title_a.get_text(strip=True) if title_a else ""
            href = title_a.get("href", "") if title_a else ""
            url_full = "https://www.jobkorea.co.kr" + href if href.startswith("/") else href
            company = company_a.get_text(strip=True) if company_a else ""
            etc_text = etc.get_text(" / ", strip=True) if etc else ""
            if not title:
                continue
            results.append({
                "site": "잡코리아",
                "title": title,
                "company": company,
                "location": etc_text,
                "experience": "1~3 년",
                "url": url_full,
                "tags": keyword,
            })
    except Exception as e:
        logger.error(f"[jobkorea] {keyword} 오류: {e}")
    return results


# ===== 필터링 / 중복제거 =====
def matches_filter(job: dict, cfg: dict) -> bool:
    text = " ".join([
        job.get("title", ""), job.get("company", ""),
        job.get("location", ""), job.get("tags", ""),
    ]).lower()
    for ex in cfg.get("exclude_keywords", []):
        if ex.lower() in text:
            return False
    # 지역 필터 (location 정보가 있을 때만)
    loc = job.get("location", "")
    if loc:
        loc_ok = any(l in loc for l in cfg.get("locations", []))
        if not loc_ok:
            # 원격 가능/전국 등의 표기도 통과
            if not any(k in text for k in ["원격", "재택", "전국"]):
                return False
    return True


def dedupe(jobs: list) -> list:
    seen = set()
    out = []
    for j in jobs:
        key = (j.get("company", "").strip(), j.get("title", "").strip())
        if key in seen or not key[1]:
            continue
        seen.add(key)
        out.append(j)
    return out


# ===== 수집 메인 =====
def collect_all(cfg: dict) -> list:
    all_jobs = []
    keywords = cfg.get("keywords", [])
    limit = int(cfg.get("max_per_site", 20))

    logger.info(f"키워드 검색 시작: {keywords}")

    for kw in keywords:
        if cfg["sites"].get("wanted"):
            logger.info(f"[원티드] 검색: {kw}")
            all_jobs.extend(fetch_wanted(kw, limit=limit))
            time.sleep(0.8)
        if cfg["sites"].get("saramin"):
            logger.info(f"[사람인] 검색: {kw}")
            all_jobs.extend(fetch_saramin(kw, limit=limit))
            time.sleep(0.8)
        if cfg["sites"].get("jobkorea"):
            logger.info(f"[잡코리아] 검색: {kw}")
            all_jobs.extend(fetch_jobkorea(kw, limit=limit))
            time.sleep(0.8)

    filtered = [j for j in all_jobs if matches_filter(j, cfg)]
    deduped = dedupe(filtered)

    logger.info(f"총 {len(all_jobs)}건 수집 → 필터링 후 {len(filtered)}건 → 중복제거 후 {len(deduped)}건")
    return deduped


# ===== 리포트 생성 =====
def to_html(jobs: list, today: str) -> str:
    by_site = {}
    for j in jobs:
        by_site.setdefault(j["site"], []).append(j)

    html = [f"""<!doctype html><html><head><meta charset="utf-8"><style>
body {{ font-family: -apple-system, 'Apple SD Gothic Neo', sans-serif; color:#222; }}
h1 {{ color:#1a73e8; }}
h2 {{ border-bottom:2px solid #eee; padding-bottom:4px; margin-top:28px; }}
table {{ border-collapse:collapse; width:100%; margin-top:8px; font-size:14px; }}
th, td {{ border:1px solid #e5e5e5; padding:8px; vertical-align:top; text-align:left; }}
th {{ background:#f5f7fb; }}
tr:nth-child(even) td {{ background:#fafbfc; }}
a {{ color:#1a73e8; text-decoration:none; }}
.small {{ color:#888; font-size:12px; }}
</style></head><body>"""]
    html.append(f"<h1>오늘의 채용공고 ({today})</h1>")
    html.append(f"<p class='small'>백엔드/풀스택 · 주니어 (1~3 년) · 서울/수도권 · 총 <b>{len(jobs)}</b> 건</p>")

    if not jobs:
        html.append("<p>오늘은 조건에 맞는 새 공고가 없어요. 내일 다시 확인할게요.</p>")

    for site, items in by_site.items():
        html.append(f"<h2>{site} ({len(items)} 건)</h2>")
        html.append("<table><tr><th>회사</th><th>공고</th><th>지역/조건</th><th>태그</th></tr>")
        for j in items:
            title_link = f"<a href='{j.get('url','')}' target='_blank'>{j.get('title','')}</a>" if j.get("url") else j.get("title", "")
            html.append(
                f"<tr><td>{j.get('company','')}</td>"
                f"<td>{title_link}</td>"
                f"<td>{j.get('location','')}</td>"
                f"<td>{j.get('tags','')}</td></tr>"
            )
        html.append("</table>")
    html.append("</body></html>")
    return "".join(html)


def to_xlsx(jobs: list, path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.info("openpyxl 미설치 - 설치 중...")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "채용공고"
    headers = ["사이트", "회사", "공고명", "지역/조건", "경력", "태그", "URL"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A73E8")
        c.alignment = Alignment(horizontal="center", vertical="center")
    widths = [10, 20, 50, 30, 10, 30, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for j in jobs:
        ws.append([
            j.get("site", ""), j.get("company", ""), j.get("title", ""),
            j.get("location", ""), j.get("experience", ""),
            j.get("tags", ""), j.get("url", ""),
        ])
    wb.save(path)


# ===== 이메일 =====
def send_email(cfg: dict, html: str, attachment_path: Path, today: str):
    em = cfg.get("email", {})
    if not em.get("enabled"):
        logger.info("[email] 비활성화됨 — 파일만 저장")
        return False
    try:
        msg = MIMEMultipart()
        msg["Subject"] = f"[채용공고 Daily] {today} · 백엔드 주니어"
        msg["From"] = em["sender"]
        msg["To"] = em["recipient"]
        msg.attach(MIMEText(html, "html", "utf-8"))
        if attachment_path and attachment_path.exists():
            with open(attachment_path, "rb") as f:
                part = MIMEApplication(f.read(), Name=attachment_path.name)
            part["Content-Disposition"] = f'attachment; filename="{attachment_path.name}"'
            msg.attach(part)
        with smtplib.SMTP(em["smtp_server"], em["smtp_port"]) as s:
            s.starttls()
            s.login(em["sender"], em["app_password"])
            s.send_message(msg)
        logger.info(f"[email] 발송 완료 → {em['recipient']}")
        return True
    except Exception as e:
        logger.error(f"[email] 발송 실패: {e}")
        traceback.print_exc()
        return False


# ===== 메인 =====
def main():
    today = datetime.now().isoformat()
    logger.info("=" * 50)
    logger.info(f"=== 채용공고 수집 시작 ({today}) ===")

    try:
        cfg = load_config()
        jobs = collect_all(cfg)
        logger.info(f"[결과] 총 {len(jobs)}건 수집")

        xlsx_path = REPORTS_DIR / f"jobs_{today}.xlsx"
        to_xlsx(jobs, xlsx_path)
        logger.info(f"[파일] 엑셀 저장: {xlsx_path}")

        html = to_html(jobs, today)
        html_path = REPORTS_DIR / f"jobs_{today}.html"
        html_path.write_text(html, encoding="utf-8")
        logger.info(f"[파일] HTML 저장: {html_path}")

        try:
            send_email(cfg, html, xlsx_path, today)
        except Exception as e:
            logger.error(f"[email] 발송 실패: {e}")
            traceback.print_exc()

        save_status(len(jobs), f"성공 - {len(jobs)}건 수집")
        return {
            "count": len(jobs),
            "xlsx": str(xlsx_path),
            "html": str(html_path),
            "status": "success",
        }
    except Exception as e:
        logger.error(f"[오류] 메인 실행 실패: {e}")
        traceback.print_exc()
        save_status(-1, f"실패 - {e}")
        return {
            "status": "error",
            "error": str(e),
        }


def get_status():
    """현재 봇 상태 조회"""
    status_file = BASE_DIR / "status.json"
    if not status_file.exists():
        print("실행 기록이 없습니다.")
        return
    with open(status_file, "r", encoding="utf-8") as f:
        status = json.load(f)
    print(json.dumps(status, ensure_ascii=False, indent=2))


def cmd_test():
    """로컬 테스트 모드"""
    print("[테스트 모드] systemd 로그 없이 바로 실행")
    logger.handlers = [console_handler]  # 파일 핸들러 제거
    return main()


def cmd_status():
    """상태 조회"""
    get_status()


def cmd_version():
    """버전 정보"""
    print("job_bot 1.0.0 - 채용공고 수집 봇 (systemd 자동화)")


def main_cli():
    """CLI 진입점"""
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == "test":
            cmd_test()
        elif cmd == "status":
            cmd_status()
        elif cmd == "version":
            cmd_version()
        else:
            print(f"사용법: python job_bot_main.py [test|status|version]")
            print("  test - 로컬 테스트 실행")
            print("  status - 실행 상태 조회")
            print("  version - 버전 정보")
    else:
        # systemd 실행 모드 (로그 파일에 기록)
        main()


if __name__ == "__main__":
    main_cli()
